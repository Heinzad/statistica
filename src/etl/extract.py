#src\etl\extract.py

"""etl 

Extract Module  
============== 

Extracts source data into a data store. 


Methods 
------- 

ingest_geospatial_file(file_path:str, table_name:str, db_path:str): 
    Read a given geospatial file and write to a given data store. Returns nothing. 

ingest_spreadsheet_table(sheet_name:str, file_path:str, skiprows:int, table_name:str, db_path:str): 
    Read a data table in a given spreadsheet and write to a given data store. Returns nothing. 

ingest_spreadsheet_range(sheet_name:str, file_path:str, skiprows:int, nrows:int, table_name:str, db_path:str): 
    Read a range of cells in a given spreadsheet and write to a given data store. Returns nothing.

ingest_spreadsheet_body(sheet_name:str, file_path:str, skiprows:int, table_name:str, db_path:str): 
    Read the body of a pivot table with hierachical headers in a given spreadsheet and write to a given data store. Returns nothing.

ingest_spreadsheet_head(sheet_name:str, file_path:str, skiprows:int, nrows:int, survey:str, dated:str, section:str, table_name:str, db_path:str='Questions'): 
    Read the body of a pivot table with hierachical headers in a given spreadsheet and write to a given data store. Returns nothing.

ingest_access_db : 
    Extract Microsoft Access database tables into a sqlite database —  
    Returns number of database records created.

_put_dataframe : 
    Store dataframes in a sqlite database — 
    Returns number of records created in database for a given dataframe, table name, and database path

_get_geospatial_file : 
    Extract a geospatial file — 
    Returns a geopandas geodataframe 

_get_spreadsheet_table : 
    Extract a data table that has headers — 
    Returns a pandas dataframe
_set_spreadsheet_table : 
    Returns a pandas dataframe

_get_spreadsheet_range : 
    Extract a range of cells without headers — 
    Returns a pandas dataframe 
_set_spreadsheet_range : 
    Returns a pandas dataframe

_get_spreadsheet_body : 
    Extract and unpivot the body of a pivot table — 
    Returns a pandas dataframe
_set_spreadsheet_body_geog : 
    Returns a pandas dataframe
_set_spreadsheet_body_head : 
    Returns a pandas dataframe

_get_spreadsheet_head : 
    Extract and unpivot hierachical headers of a pivot table — 
    Returns a pandas dataframe 
_set_spreadsheet_head : 
    Returns a pandas dataframe


History
-------

20250427 -- Add helper functions
20250428 -- prefix helper function names with an underscore to indicate they are private
20250429 -- Add pipelines to integrate helper functions in a sequence to extract from source file to a sink data store
20250430 -- decompose helper functions into _get and _set, pipe with _put in ingest methods.
20250501 -- Add simple module logging

"""
import os
import functools
import logging
import pandas as pd
import geopandas as gpd
from openpyxl.utils import get_column_letter 
from typing import Dict, List
from sqlalchemy.types import NVARCHAR
from datetime import datetime, timezone
from dotenv import load_dotenv
from src.db.connect import connect_mdb, connect_db 

load_dotenv()

PREFIX1 = 'count_'
PREFIX2 = 'geog_'   

# Logging setup 

LOGDIR = os.getenv('LOGDIR')
FORMAT='%(asctime)s.%(msecs)03d %(filename)s %(lineno)s %(levelname)s | %(message)s'
DATEFORMAT='%Y-%m-%d %H:%M:%S' 
BATCH_ID = int(datetime.now(timezone.utc).timestamp() * 1000000)

def get_logger(log_file_name):
    """ Creates and returns logger object """
    log_path = LOGDIR + log_file_name.replace(chr(46),chr(95)) + '.log'
    logger = logging.getLogger(__name__)
    formatter = logging.Formatter(fmt=FORMAT,datefmt=DATEFORMAT)
    logger.handlers.clear()
    file_handler = logging.FileHandler(log_path, 'a') 
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.setLevel(logging.DEBUG)
    logger.propagate=False
    return logger

def prettify_return_type(class_type: str)->str: 
    result_type1 = class_type[1:-1].split(' ')
    result_type2 = result_type1[1]
    return result_type2[1:-1]
 
def log_decorator(_func=None):
    def log_decorator_info(func):
        @functools.wraps(func)
        def log_decorator_wrapper(*args, **kwargs): 
            msg = {'BATCH': BATCH_ID} 
            try: 
                logger_obj = get_logger(log_file_name=__name__) 
                msg.update({ 'FUNCTION': func.__name__})
                p_args = list()
                for a in args: 
                    if isinstance(a, pd.DataFrame):  
                        p_args.append("dataframe=DataFrame")
                    elif isinstance(a, gpd.GeoDataFrame): 
                        p_args.append("dataframe=GeoDataFrame")
                    else: 
                        p_args.append(repr(a))
                p_kwargs = [f"{k}={v!r}" for k, v in kwargs.items()]
                params = ", ".join(p_args + p_kwargs)
                msg.update({'PARAMS': params })
            except Exception as e:
                msg.update({'ERROR': e})
                logger_obj.error(str(msg)) 
                raise 
            try:
                results = func(*args, **kwargs)
                return_type = prettify_return_type(str(results.__class__))
                returns = {'type': return_type }
                items_returned = list()
                item_count = 1
                is_iterable = False
                if return_type in ('list', 'tuple'): 
                    is_iterable = True 
                if is_iterable: 
                    item_count = len(results)
                for i in range(item_count): 
                    if is_iterable: 
                        result = results[i]
                        return_type = prettify_return_type(str(result.__class__))
                    else: 
                        result = results
                    if return_type in ('str', 'int', 'float'):
                        returns.update({'rows': result})
                    else: 
                        try: 
                            items_returned.append({'type': return_type, 'rows': len(result), 'cols': len(result.columns)})
                        except: 
                            returns.update({'rows': len(result)})
                if is_iterable: 
                    returns.update({'items': items_returned})
                msg.update({ 'RETURN': returns }) 
            except Exception as e:
                msg.update({'ERROR': e})
                logger_obj.error(str(msg))
                raise
            logger_obj.info(str(msg)) 
            return results
        return log_decorator_wrapper
    if _func is None:
        return log_decorator_info
    else:
        return log_decorator_info(_func)


# pandas pipelines

def ingest_geospatial_file(file_path:str, table_name:str, db_path:str): 
    """Read a given geospatial file and write to a given data store. Returns nothing. """
    _geospatial = _get_geospatial_file(file_path=file_path)
    _geospatial.pipe(_set_geospatial_file).pipe(_put_dataframe, table_name=table_name, db_path=db_path)

def ingest_spreadsheet_table(sheet_name:str, file_path:str, skiprows:int, table_name:str, db_path:str): 
    """Read a data table in a given spreadsheet and write to a given data store. Returns nothing. """
    _table = _get_spreadsheet_table(sheet_name=sheet_name, file_path=file_path, skiprows=skiprows)
    _table.pipe(_set_spreadsheet_table).pipe(_put_dataframe, table_name=table_name, db_path=db_path)

def ingest_spreadsheet_range(sheet_name:str, file_path:str, skiprows:int, nrows:int, column_names:List, table_name:str, db_path:str): 
    """Read a range of cells in a given spreadsheet and write to a given data store. Returns nothing. """
    _range = _get_spreadsheet_range(sheet_name=sheet_name, file_path=file_path, skiprows=skiprows, nrows=nrows)
    _range.pipe(_set_spreadsheet_range, column_names=column_names).pipe(_put_dataframe, table_name=table_name, db_path=db_path)

def ingest_spreadsheet_body(sheet_name:str, file_path:str, skiprows:int, table_name:str, db_path:str): 
    """Read the body of a pivot table with hierachical headers in a given spreadsheet and write to a given data store. Returns nothing. """
    _body = _get_spreadsheet_body(sheet_name=sheet_name, file_path=file_path, skiprows=skiprows)
    _body.pipe(_set_spreadsheet_body_count, table_name=table_name).pipe(_put_dataframe, table_name=PREFIX1+table_name, db_path=db_path)
    _body.pipe(_set_spreadsheet_body_geog, table_name=table_name).pipe(_put_dataframe, table_name=PREFIX2+table_name, db_path=db_path)

def ingest_spreadsheet_head(sheet_name:str, file_path:str, skiprows:int, nrows:int, survey:str, dated:str, section:str, table_name:str, db_path:str='Questions'): 
    """Read the body of a pivot table with hierachical headers in a given spreadsheet and write to a given data store. Returns nothing. """
    _head = _get_spreadsheet_head(sheet_name=sheet_name, file_path=file_path, skiprows=skiprows, nrows=nrows)
    _head.pipe(_set_spreadsheet_head, survey=survey, dated=dated, section=section, table_name=table_name).pipe(_put_dataframe, table_name=table_name, db_path=db_path)
        

@log_decorator
def ingest_access_db(mdb_path:str, db_path:str)->Dict: 
    """Read Microsoft Access database tables and write into a sqlite database — 
    Returns dictionary of number of rows inserted per table. 
    
    Parameters
    ----------
    mdb_path : str 
        Absolute path to a Microsoft Access database file
    db_path : str 
        Absolute path to database file 
    
    Returns
    -------
    dict 
        { table name : row count }
    
    Example
    -------
    >>> src = store_access_db(
    ... mdb_path = 'C:/Users/Public/Documents/CensusData.mdb',
    ... db_path = 'C:/Users/Public/Documents/test_db.sqlite'
    ... )
    >>> [print(k,v) for k,v in extract]
    
    """ 
    table_name = str()
    row_count = int()
    results = list()
    mdb_conn = connect_mdb(mdb_path=mdb_path)
    mdb_cursor = mdb_conn.cursor()
    mdb_tables = mdb_cursor.tables(tableType='TABLE')
    for tbl in mdb_tables: 
        row_count = 0 
        table_name = tbl.table_name 
        dataframe = pd.read_sql(f"""SELECT * FROM "{table_name}";""", mdb_conn)
        row_count = _put_dataframe(dataframe=dataframe, table_name=table_name, db_path=db_path)
        results.append({table_name: row_count})
    return results



# helper functions 

@log_decorator
def _put_dataframe(dataframe: pd.DataFrame, table_name:str, db_path:str) -> (int | None): 
    """Store dataframes in a sqlite database — 
    Returns number of records created on saving a given DataFrame to a database table with a given name. 
    
    Parameters
    ----------
    dataframe : pandas.DataFrame
        The DataFrame to be stored as a database table. 
    table_name : str
        Name of the table to be created. 
    db_path : str 
        Absolute path to database file 

    Returns
    -------
    int or None
        Number of rows affected by to_sql (rowcount), otherwise None.  
    
    Example
    -------
    >>> file_path = 'C:/Users/Public/Documents/AU1996.ZIP'
    >>> dataframe = _get_geospatial_file(file_path=file_path, db_path=db_path, table_alias=table_alias)
    >>> table_name = 'AU1996'
    >>> db_path = 'C:/Users/Public/Documents/test_db.sqlite'
    >>> result = _put_dataframe(dataframe=dataframe, table_name=table_name, db_path)
    >>> print(result)

    """ 
    db_conn = connect_db(db_path=db_path)
    return dataframe.to_sql(
        name= table_name,
        con=db_conn,
        schema=None, # default schema
        if_exists='replace', 
        index_label='_id', 
        chunksize=1048576, 
        dtype=NVARCHAR # import all values as text
    ) 


@log_decorator
def _get_geospatial_file(file_path:str) -> gpd.GeoDataFrame: 
    """Extract a geospatial file — 
    Returns a GeoDataFrame for a given geospatial file. 
    
    Parameters
    ----------
    file_path : str
        Absolute path to the file

    Returns
    -------
    geopandas.GeoDataFrame 

    Example
    -------
    >>> src = _get_geospatial_file(file_path = 'zip:///Users/Public/Documents/AU1996.ZIP') 
    >>> print(src)
    """
    return gpd.read_file(file_path)   

@log_decorator
def _set_geospatial_file(dataframe:gpd.GeoDataFrame ) -> gpd.GeoDataFrame:
    """Returns a GeoDataFrame with geometries converted to Well-Known-Text 
    
    Parameters
    ----------
    dataframe : geopandas.GeoDataFrame

    Returns
    -------
    geopandas.GeoDataFrame 
    """
    return dataframe.to_wkt()


@log_decorator
def _get_spreadsheet_table(sheet_name:str, file_path:str, skiprows:int) -> pd.DataFrame: 
    """Extract a data table that has headers — 
    Returns a pandas DataFrame for a data table in a given excel workbook
    
    Parameters
    ----------
    sheet_name : str 
        Name of the spreadsheet to be extracted from an excel workbook. 
    file_path : str
        Absolute path to the Excel workbook containing the census results. 
    skiprows : int
        Number of rows to skip to beginning of data table. 

    Returns
    -------
    pandas.DataFrame

    Example
    ------- 
    >>> src = _get_spreadsheet_table(
    ... sheet_name = 'Geographic Key',
    ... file_path = "C:/Users/Public/Documents/2013-mb-dataset-Total-New-Zealand-individual-part-1.xlsx",
    ... skiprows = 2
    ... )
    >>> print(src)
    
    """ 
    return pd.read_excel(
        io = pd.ExcelFile(file_path), 
        sheet_name=sheet_name, 
        skiprows=skiprows, 
        header=0, 
        dtype='object', 
        engine='openpyxl'
    )

@log_decorator
def _set_spreadsheet_table(dataframe:pd.DataFrame)->pd.DataFrame: 
    """Returns a cleansed dataframe"""
    tabled = dataframe.dropna() #inplace = True
    return tabled


@log_decorator
def _get_spreadsheet_range(sheet_name:str, file_path:str, skiprows:int, nrows:int) -> pd.DataFrame: 
    """Extract a range of cells without headers — 
    Returns a pandas DataFrame for a range in a given excel workbook
    
    Parameters
    ----------
    sheet_name : str 
        Name of the spreadsheet to be extracted from an excel workbook. 
    file_path : str
        Absolute path to the Excel workbook containing the census results. 
    skiprows : int
        Number of rows to skip to beginning of data table. 
    nrows : int
        Number of rows to extract from data table headers.

    Returns
    -------
    pandas.DataFrame
        Enables pipelining

    Example
    ------- 
    >>> src = _get_spreadsheet_range(
    ... sheet_name = 'Footnotes and symbols',
    ... file_path = "C:/Users/Public/Documents/2013-mb-dataset-Total-New-Zealand-individual-part-1.xlsx",
    ... skiprows = 12,
    ... nrows = 14
    ... )
    >>> print(src)
    
    """ 
    return pd.read_excel(
        io = pd.ExcelFile(file_path), 
        sheet_name=sheet_name, 
        skiprows=skiprows, 
        nrows=nrows, 
        dtype='object', 
        engine='openpyxl'
    )

@log_decorator
def _set_spreadsheet_range(dataframe:pd.DataFrame, column_names: List)->pd.DataFrame:
    """
    Returns a cleansed dataframe sourced from a range of cells in a spreadsheet

    Parameters
    ----------
    pandas.DataFrame
        Enables pipelining
    
    Returns
    ----------
    pandas.DataFrame
        Enables pipelining
    """
    for col in dataframe: 
        c = list(dataframe.columns).index(col)
        alpha = get_column_letter( int(c) + 1 ) #avoid zero 
        dataframe.rename(columns={col: alpha}, inplace = True) 
    dataframe.columns = column_names 
    ranged = dataframe.dropna() 
    return ranged


@log_decorator
def _get_spreadsheet_body(sheet_name:str, file_path:str, skiprows:int) -> pd.DataFrame:
    """Extract and unpivot the body of a pivot table into a long dataframe — 
    Returns a tuple of pandas DataFrames for geographies and counts respectively 
    from the body of a pivot table in a given excel workbook
    
    Parameters
    ----------
    sheet_name : str 
        Name of the spreadsheet to be extracted from an excel workbook. 
    file_path : str
        Absolute path to the Excel workbook containing the census results. 
    skiprows : int
        Number of rows to skip to beginning of data table. 
    table_name : str 
        Name to be used in database table. 
    
    Returns
    ----------
    pandas.DataFrame
        Enables pipelining
    
    Example
    -------
    >>> src = _get_spreadsheet_body(
    ... sheet_name = '1 Meshblock',
    ... file_path = "C:/Users/Public/Documents/2013-mb-dataset-Total-New-Zealand-individual-part-1.xlsx",
    ... skiprows = 10,
    ... table_name = 'MeshBlock'
    ... )
    >>> print(src)

    """
    return pd.read_excel(
        io = pd.ExcelFile(file_path), 
        sheet_name=sheet_name, 
        skiprows=skiprows, 
        header=None,
        dtype='object', 
        engine='openpyxl'
    ) 

@log_decorator
def _set_spreadsheet_body_geog(dataframe:pd.DataFrame, table_name:str)->pd.DataFrame: 
    """Returns a cleansed geographies dataframe reshaped from wide to long
    
    Parameters
    ----------
    pandas.DataFrame
        Enables pipelining
    
    """
    _name = table_name.lower()
    for col in dataframe: 
        c = list(dataframe).index(col)
        alpha = get_column_letter( int(c) + 1 ) 
        dataframe.rename(columns={col: alpha}, inplace = True) 
    dataframe['A'] = dataframe['A'].fillna('00') 
    if _name == 'meshblock': 
        dfg = dataframe[['A']] 
        dfg.columns = [f'{_name}_code']
    else: 
        dfg = dataframe[['A', 'B']] 
        dfg.columns = [f'{_name}_code', f'{_name}_description'] 
    return dfg 

@log_decorator
def _set_spreadsheet_body_count(dataframe:pd.DataFrame, table_name:str)->pd.DataFrame: 
    """Returns a cleansed counts dataframe reshaped from wide to long
    
    Parameters
    ----------
    pandas.DataFrame
        Enables pipelining

    Returns
    ----------
    pandas.DataFrame
        Enables pipelining
    
    """
    _name = table_name.lower()
    for col in dataframe: 
        c = list(dataframe).index(col)
        alpha = get_column_letter( int(c) + 1 ) 
        dataframe.rename(columns={col: alpha}, inplace = True) 
    dataframe['A'] = dataframe['A'].fillna('00') 
    if _name == 'meshblock': 
        dfc0 = dataframe 
    else: 
        dfc0 = dataframe.drop(['B'], axis=1) 
    dfc1 = dfc0.set_index(['A']).stack() 
    dfc2 = dfc1.to_frame() 
    dfc = dfc2.reset_index() 
    dfc.columns = [f'{_name}_code', 'question_code', 'response_count'] 
    return dfc  


@log_decorator
def _get_spreadsheet_head(sheet_name:str, file_path:str, skiprows:int, nrows:int)->pd.DataFrame: 
    """Extract and unpivot hierarchical headers of a pivot table — 
    Returns a pandas dataframe from unpivoted multi-line headings in the head of a pivot table in a 
    given excel workbook. 
    
    Parameters
    ----------
    sheet_name : str 
        Name of the spreadsheet to be extracted from an excel workbook. 
    file_path : str
        Absolute path to the Excel workbook containing the census results. 
    skiprows : int
        Number of rows to skip to beginning of data table. 
    nrows : int
        Number of rows to extract from data table headers. 
     
    
    Returns
    -------
    pandas.DataFrame
        Enables pipelining

    Example
    -------
    >>> src = _get_spreadsheet_head(
    ... sheet_name = '5 Regional Council Area',
    ... file_path = "C:/Users/Public/Documents/2013-mb-dataset-Total-New-Zealand-individual-part-1.xlsx",
    ... skiprows = 8,
    ... nrows = 2
    ... )
    >>> print(src)
    """
    return pd.read_excel(
        io = pd.ExcelFile(file_path), 
        sheet_name=sheet_name, 
        skiprows=skiprows,
        header=None,
        nrows=nrows,
        dtype='object',
        engine='openpyxl'
    )

@log_decorator
def _set_spreadsheet_head(dataframe:pd.DataFrame, survey:str, dated:str, section:str, table_name:str='Questions')->pd.DataFrame:
    """Returns a cleansed question dataframe reshaped from wide to long
    
    Parameters
    ----------
    pandas.DataFrame
        Enables pipelining
    survey : str
        The name of the survey - e.g. 'Census'. 
    dated : str 
        The date of the survey - e.g. '2013'.
    section : str
        The section of the survey being extracted - e.g. 'Individual part 1'

    Returns
    -------
    pandas.DataFrame
        Enables pipelining

    Example
    -------
    >>> stg = _set_spreadsheet_head(
    ... dataframe = src,
    ... survey = 'Census',
    ... dated = '2013',
    ... section = 'individual part 1',
    ... table_name = 'Questions'
    ... )

    """ 
    for col in dataframe.columns: 
        c = list(dataframe.columns).index(col)
        alpha = get_column_letter( int(c) + 1 )
        dataframe.rename(columns={col: alpha}, inplace = True) 
    pvt0 = dataframe.transpose() 
    pvt0.reset_index() 
    pvt0.columns = ['question_text', 'question_text2']  
    pvt0['question_text'] = pvt0['question_text'].ffill(axis=0) 
    pvt0['question_code'] = pvt0.index 
    pvt0.reset_index(drop=True, inplace=True)
    pvt0['survey_name'] = survey
    pvt0['survey_date'] = dated 
    pvt0['survey_section'] = section
    pvt = pvt0.loc[:,['question_code', 'question_text', 'question_text2', 'survey_name', 'survey_date', 'survey_section']]
    return pvt 

    